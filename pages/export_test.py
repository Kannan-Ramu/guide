from django.http import HttpResponse
from openpyxl import Workbook
from .models import Team


def export_to_excel(request):
    # apply filters as necessary
    queryset = Team.objects.filter(guide_approved=True)

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add column headings to the worksheet
    fields = Team._meta.get_fields()
    col = 1
    for field in fields:
        ws.cell(row=1, column=col, value=field.name.capitalize())
        col += 1

    # Add data to the worksheet
    row = 2
    for obj in queryset:
        col = 1
        for field in fields:
            value = getattr(obj, field.name)
            ws.cell(row=row, column=col, value=value)
            col += 1
        row += 1

    # Create an HTTP response with the Excel file as the content
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
    wb.save(response)

    return response
